import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import json

# Page configuration
st.set_page_config(
    page_title="Security Rota AI", 
    layout="wide", 
    page_icon="📅",
    initial_sidebar_state="expanded"
)

# Initialize session state
def init_session_state():
    if 'employees' not in st.session_state:
        st.session_state.employees = [
            {
                'id': 1,
                'name': 'John Smith',
                'phone': '07700 123456',
                'email': 'john.smith@email.com',  # Added email
                'postcode': 'LE1 1AA',
                'sia_license': 'SIA123456',
                'max_hours': 48,
                'availability': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
                'willing_24hr': True
            },
            {
                'id': 2,
                'name': 'Sarah Wilson',
                'phone': '07700 789012',
                'email': 'sarah.wilson@email.com',  # Added email
                'postcode': 'NN18 8BB',
                'sia_license': 'SIA789012',
                'max_hours': 40,
                'availability': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
                'willing_24hr': False
            }
        ]
    
    if 'sites' not in st.session_state:
        st.session_state.sites = []
    
    if 'schedules' not in st.session_state:
        st.session_state.schedules = {}
    
    if 'current_schedule' not in st.session_state:
        st.session_state.current_schedule = None
    
    if 'alerts' not in st.session_state:
        st.session_state.alerts = []
    
    if 'next_employee_id' not in st.session_state:
        st.session_state.next_employee_id = 3
    
    if 'next_site_id' not in st.session_state:
        st.session_state.next_site_id = 3

# Helper functions
def calculate_shift_hours(start_time, end_time):
    """Calculate hours between two times (handles overnight shifts)"""
    start_h, start_m = map(int, start_time.split(':'))
    end_h, end_m = map(int, end_time.split(':'))
    
    start_minutes = start_h * 60 + start_m
    end_minutes = end_h * 60 + end_m
    
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    
    return (end_minutes - start_minutes) / 60

def estimate_distance(postcode1, postcode2):
    """Simplified distance estimation"""
    if not postcode1 or not postcode2:
        return 999
    
    area1 = ''.join(filter(str.isalpha, postcode1[:3])).upper()
    area2 = ''.join(filter(str.isalpha, postcode2[:3])).upper()
    
    if area1 == area2:
        return 5
    elif area1[0] == area2[0]:
        return 25
    else:
        return 50

# Employee Management
def manage_employees():
    st.title("👥 Manage Employees")
    
    with st.expander("➕ Add New Employee", expanded=False):
        with st.form("add_employee"):
            col1, col2 = st.columns(2)
            
            with col1:
                name = st.text_input("Full Name*")
                phone = st.text_input("Phone Number")
                postcode = st.text_input("Home Postcode*")
                email = st.text_input("Email Address*")  # Added email field
            
            with col2:
                sia_license = st.text_input("SIA License Number")
                max_hours = st.number_input("Max Weekly Hours", min_value=1, max_value=60, value=48)
                willing_24hr = st.checkbox("Willing to work 24-hour shifts")
            
            st.write("**Available Days:**")
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            availability = st.multiselect("Select available days", days, default=days)
            
            submitted = st.form_submit_button("Add Employee")
            
            if submitted:
                if name and postcode and email:  # Check if the essential fields are filled
                    new_emp = {
                        'id': st.session_state.next_employee_id,
                        'name': name,
                        'phone': phone,
                        'email': email,  # Store the email
                        'postcode': postcode,
                        'sia_license': sia_license,
                        'max_hours': max_hours,
                        'availability': availability,
                        'willing_24hr': willing_24hr
                    }
                    st.session_state.employees.append(new_emp)
                    st.session_state.next_employee_id += 1
                    st.success(f"✅ Added {name} successfully!")
                    st.rerun()
                else:
                    st.error("Please fill in Name, Postcode, and Email Address.")
    
# Site Management
def manage_sites():
    st.title("📍 Manage Sites")
    
    # ADD NEW SITE SECTION
    with st.expander("➕ Add New Site", expanded=False):
        with st.form("add_site"):
            col1, col2 = st.columns(2)
            
            with col1:
                site_name = st.text_input("Site Name*")
                client = st.selectbox("Client", ["Taz", "Servo", "Ayam"])
                postcode = st.text_input("Postcode*")
                guards = st.number_input("Guards Required", min_value=1, max_value=10, value=1)
            
            with col2:
                shift_start = st.time_input("Shift Start")
                shift_end = st.time_input("Shift End")
            
            st.write("**Operating Days:**")
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            operation_days = st.multiselect("Select operating days", days, default=days)
            
            # Weekend shift dynamic input
            st.write("**Weekend Shifts:**")
            weekend_shifts = st.checkbox("Enable Weekend Shifts")
            
            weekend_guards = None
            shift_type = None
            if weekend_shifts:
                weekend_guards = st.number_input("How many guards required for weekends?", min_value=1, max_value=3, value=1)
                shift_type = st.radio("What type of shifts on weekends?", ['Day Shift', 'Night Shift', 'Day & Night'])
            
            submitted = st.form_submit_button("Add Site")
            
            if submitted:
                if site_name and postcode and shift_start and shift_end:
                    new_site = {
                        'id': st.session_state.next_site_id,
                        'name': site_name,
                        'client': client,
                        'postcode': postcode,
                        'guards_required': guards,
                        'shift_start': shift_start.strftime("%H:%M"),
                        'shift_end': shift_end.strftime("%H:%M"),
                        'weekend_shifts_enabled': weekend_shifts,
                        'weekend_guards': weekend_guards,
                        'shift_type': shift_type,
                        'days_operation': operation_days
                    }
                    st.session_state.sites.append(new_site)
                    st.session_state.next_site_id += 1
                    st.success(f"✅ Added {site_name} successfully!")
                    st.rerun()
                else:
                    st.error("Please fill in all required fields")
    
    # DISPLAY EXISTING SITES SECTION (THIS WAS MISSING!)
    st.subheader("Current Sites")
    
    if st.session_state.sites:
        for site in st.session_state.sites:
            with st.expander(f"🏢 {site['name']} ({site['client']}) - {site['postcode']}"):
                col1, col2, col3 = st.columns([2, 2, 1])
                
                with col1:
                    st.write(f"**Guards Required:** {site['guards_required']}")
                    st.write(f"**Shift:** {site['shift_start']} - {site['shift_end']}")
                
                with col2:
                    st.write(f"**Operating Days:** {', '.join(site['days_operation'])}")
                    hours = calculate_shift_hours(site['shift_start'], site['shift_end'])
                    st.write(f"**Shift Duration:** {hours:.1f} hours")
                    
                    # Display weekend shift info if enabled
                    if site.get('weekend_shifts_enabled'):
                        st.write(f"**Weekend Guards:** {site.get('weekend_guards', 'N/A')}")
                        st.write(f"**Weekend Shift Type:** {site.get('shift_type', 'N/A')}")
                
                with col3:
                    if st.button("🗑️ Delete", key=f"del_site_{site['id']}"):
                        st.session_state.sites = [s for s in st.session_state.sites if s['id'] != site['id']]
                        st.success("Deleted!")
                        st.rerun()
    else:
        st.info("No sites added yet. Add your first site above!")
# Scheduling Logic and Other Functions...
# Continue your existing logic with the necessary updates for weekend shifts, guards, and more.

def export_to_excel(schedule, employees, sites, alerts, unassigned, opportunities, week_start):
    """Generate comprehensive Excel workbook"""
    wb = Workbook()
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    # Sheet 1: Weekly Schedule Grid
    ws1 = wb.active
    ws1.title = "Weekly Schedule"
    
    headers = ['Employee'] + days + ['Total Hours']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row, emp in enumerate(employees, 2):
        ws1.cell(row, 1, emp['name']).border = border
        
        total_hours = 0
        for col, day in enumerate(days, 2):
            if emp['id'] in schedule and day in schedule[emp['id']]:
                shifts = schedule[emp['id']][day]
                if shifts:
                    shift = shifts[0]
                    cell_value = f"{shift['site_name']}\n{shift['start']}-{shift['end']}"
                    total_hours += shift['hours']
                else:
                    cell_value = "OFF"
            else:
                cell_value = "OFF"
            
            cell = ws1.cell(row, col, cell_value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        cell = ws1.cell(row, len(days) + 2, f"{total_hours:.1f}")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    ws1.column_dimensions['A'].width = 20
    for col in range(2, len(days) + 2):
        ws1.column_dimensions[chr(64 + col)].width = 18
    
    # Sheet 2: Site Coverage View
    ws2 = wb.create_sheet("Site Coverage")
    
    headers = ['Site', 'Postcode'] + days
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row, site in enumerate(sites, 2):
        ws2.cell(row, 1, site['name']).border = border
        ws2.cell(row, 2, site['postcode']).border = border
        
        for col, day in enumerate(days, 3):
            assigned_guards = []
            for emp_id, week_schedule in schedule.items():
                if day in week_schedule:
                    for shift in week_schedule[day]:
                        if shift['site_id'] == site['id']:
                            emp = next((e for e in employees if e['id'] == emp_id), None)
                            if emp:
                                assigned_guards.append(emp['name'])
            
            cell_value = "\n".join(assigned_guards) if assigned_guards else "UNASSIGNED"
            cell = ws2.cell(row, col, cell_value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            if not assigned_guards:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12
    
    # Sheet 3: Alerts & Issues
    ws3 = wb.create_sheet("Alerts & Issues")
    
    ws3.cell(1, 1, "Alert Type").font = header_font
    ws3.cell(1, 2, "Message").font = header_font
    
    for row, alert in enumerate(alerts, 2):
        ws3.cell(row, 1, alert['type'].upper())
        ws3.cell(row, 2, alert['message'])
        
        if alert['type'] == 'error':
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif alert['type'] == 'warning':
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        ws3.cell(row, 1).fill = fill
        ws3.cell(row, 2).fill = fill
    
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 70
    
    # Sheet 4: 24-Hour Opportunities
    ws4 = wb.create_sheet("24hr Opportunities")
    
    headers = ['Employee', 'Days', 'Site 1', 'Site 2', 'Distance (miles)']
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, opp in enumerate(opportunities, 2):
        ws4.cell(row, 1, opp['employee'])
        ws4.cell(row, 2, opp['day'])
        ws4.cell(row, 3, opp['site1'])
        ws4.cell(row, 4, opp['site2'])
        ws4.cell(row, 5, f"{opp['distance']:.1f}")
    
    # Sheet 5: Summary Statistics
    ws5 = wb.create_sheet("Summary")
    
    ws5.cell(1, 1, f"ROTA SUMMARY - Week of {week_start}").font = Font(bold=True, size=14)
    
    row = 3
    ws5.cell(row, 1, "Total Employees:").font = Font(bold=True)
    ws5.cell(row, 2, len(employees))
    
    row += 1
    ws5.cell(row, 1, "Total Sites:").font = Font(bold=True)
    ws5.cell(row, 2, len(sites))
    
    row += 1
    ws5.cell(row, 1, "Unassigned Shifts:").font = Font(bold=True)
    ws5.cell(row, 2, len(unassigned))
    
    row += 1
    ws5.cell(row, 1, "24-Hour Opportunities:").font = Font(bold=True)
    ws5.cell(row, 2, len(opportunities))
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# Main application
def main():
    init_session_state()
    
    # Sidebar
    st.sidebar.title("📅 Security Rota AI")
    st.sidebar.markdown("---")
    
    page = st.sidebar.radio(
        "Navigation",
        ["Dashboard", "Manage Employees", "Manage Sites", "Generate Schedule", "View Schedule"],
        label_visibility="collapsed"
    )
    
    st.sidebar.markdown("---")
    st.sidebar.info(f"👥 {len(st.session_state.employees)} Employees\n\n📍 {len(st.session_state.sites)} Sites")
    
    if page == "Dashboard":
        show_dashboard()
    elif page == "Manage Employees":
        manage_employees()
    elif page == "Manage Sites":
        manage_sites()
    elif page == "Generate Schedule":
        generate_schedule_page()
    elif page == "View Schedule":
        view_schedule()

def show_dashboard():
    st.title("📊 Dashboard")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Active Employees", len(st.session_state.employees))
    
    with col2:
        st.metric("Active Sites", len(st.session_state.sites))
    
    with col3:
        st.metric("Schedules Created", len(st.session_state.schedules))
    
    with col4:
        st.metric("System Status", "✅ Online")
    
    st.markdown("---")
    
    st.subheader("🚀 Quick Start Guide")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("### 1️⃣ Add Employees")
        st.write("Go to **Manage Employees** to add your security guards with their details and availability.")
        if st.button("→ Manage Employees", key="dash_emp"):
            st.session_state.page = "Manage Employees"
            st.rerun()
    
    with col2:
        st.markdown("### 2️⃣ Add Sites")
        st.write("Go to **Manage Sites** to add client locations with shift requirements.")
        if st.button("→ Manage Sites", key="dash_sites"):
            st.session_state.page = "Manage Sites"
            st.rerun()
    
    with col3:
        st.markdown("### 3️⃣ Generate Rota")
        st.write("Use **Generate Schedule** to create AI-optimized weekly schedules.")
        if st.button("→ Generate Schedule", key="dash_gen"):
            st.session_state.page = "Generate Schedule"
            st.rerun()
    
    st.markdown("---")
    
    st.subheader("📈 Current Data Overview")
    
    if st.session_state.employees:
        st.write("**Recent Employees:**")
        df_emp = pd.DataFrame(st.session_state.employees)
        st.dataframe(df_emp[['name', 'postcode', 'max_hours']], use_container_width=True)
    
    if st.session_state.sites:
        st.write("**Recent Sites:**")
        df_sites = pd.DataFrame(st.session_state.sites)
        st.dataframe(df_sites[['name', 'client', 'postcode', 'guards_required']], use_container_width=True)

def manage_employees():
    st.title("👥 Manage Employees")
    
    with st.expander("➕ Add New Employee", expanded=False):
        with st.form("add_employee"):
            col1, col2 = st.columns(2)
            
            with col1:
                name = st.text_input("Full Name*")
                phone = st.text_input("Phone Number")
                postcode = st.text_input("Home Postcode*")
            
            with col2:
                sia_license = st.text_input("SIA License Number")
                max_hours = st.number_input("Max Weekly Hours", min_value=1, max_value=60, value=48)
                willing_24hr = st.checkbox("Willing to work 24-hour shifts")
            
            st.write("**Available Days:**")
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            availability = st.multiselect("Select available days", days, default=days)
            
            submitted = st.form_submit_button("Add Employee")
            
            if submitted:
                if name and postcode:
                    new_emp = {
                        'id': st.session_state.next_employee_id,
                        'name': name,
                        'phone': phone,
                        'postcode': postcode,
                        'sia_license': sia_license,
                        'max_hours': max_hours,
                        'availability': availability,
                        'willing_24hr': willing_24hr
                    }
                    st.session_state.employees.append(new_emp)
                    st.session_state.next_employee_id += 1
                    st.success(f"✅ Added {name} successfully!")
                    st.rerun()
                else:
                    st.error("Please fill in Name and Postcode")
    
    st.subheader("Current Employees")
    
    if st.session_state.employees:
        for emp in st.session_state.employees:
            with st.expander(f"👤 {emp['name']} - {emp['postcode']}"):
                col1, col2, col3 = st.columns([2, 2, 1])
                
                with col1:
                    st.write(f"**Phone:** {emp['phone'] or 'N/A'}")
                    st.write(f"**SIA License:** {emp['sia_license'] or 'N/A'}")
                
                with col2:
                    st.write(f"**Max Hours:** {emp['max_hours']}")
                    st.write(f"**Available:** {', '.join(emp['availability'])}")
                    st.write(f"**24hr Shifts:** {'✅ Yes' if emp['willing_24hr'] else '❌ No'}")
                
                with col3:
                    if st.button("🗑️ Delete", key=f"del_emp_{emp['id']}"):
                        st.session_state.employees = [e for e in st.session_state.employees if e['id'] != emp['id']]
                        st.success("Deleted!")
                        st.rerun()
    else:
        st.info("No employees added yet. Add your first employee above!")

def manage_sites():
    st.title("📍 Manage Sites")
    
    with st.expander("➕ Add New Site", expanded=False):
        with st.form("add_site"):
            col1, col2 = st.columns(2)
            
            with col1:
                site_name = st.text_input("Site Name*")
                client = st.selectbox("Client", ["Taz", "Servo", "Ayam"])
                postcode = st.text_input("Postcode*")
                guards = st.number_input("Guards Required", min_value=1, max_value=10, value=1)
            
            with col2:
                shift_start = st.time_input("Shift Start")
                shift_end = st.time_input("Shift End")
            
            st.write("**Operating Days:**")
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            operation_days = st.multiselect("Select operating days", days, default=days)
            
            # Weekend shift dynamic input
            st.write("**Weekend Shifts**:")
            weekend_shifts = st.checkbox("Enable Weekend Shifts")
            
            weekend_guards = None
            shift_type = None
            if weekend_shifts:
                weekend_guards = st.number_input("How many guards required for weekends?", min_value=1, max_value=3)
                shift_type = st.radio("What type of shifts on weekends?", ['Day Shift', 'Night Shift', 'Day & Night'])
            
            submitted = st.form_submit_button("Add Site")
            
            if submitted:
                if site_name and postcode and shift_start and shift_end:
                    new_site = {
                        'id': st.session_state.next_site_id,
                        'name': site_name,
                        'client': client,
                        'postcode': postcode,
                        'guards_required': guards,
                        'shift_start': shift_start.strftime("%H:%M"),
                        'shift_end': shift_end.strftime("%H:%M"),
                        'weekend_shifts_enabled': weekend_shifts,  # Store the weekend shift option
                        'weekend_guards': weekend_guards,  # Store how many guards
                        'shift_type': shift_type,  # Store shift type (Day/Night/Day & Night)
                        'days_operation': operation_days
                    }
                    st.session_state.sites.append(new_site)
                    st.session_state.next_site_id += 1
                    st.success(f"✅ Added {site_name} successfully!")
                    st.rerun()
                else:
                    st.error("Please fill in all required fields")

def generate_schedule_page():
    st.title("🔄 Generate Schedule")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        week_start = st.date_input(
            "Week Start Date (Monday)",
            value=datetime.now().date()
        )
    
    with col2:
        st.write("")
        st.write("")
        generate_btn = st.button("🚀 Generate Schedule", type="primary", use_container_width=True)
    
    if generate_btn:
        if not st.session_state.employees:
            st.error("❌ Please add employees first!")
            return
        
        if not st.session_state.sites:
            st.error("❌ Please add sites first!")
            return
        
        with st.spinner("Generating optimal schedule..."):
            generator = ScheduleGenerator(week_start)
            schedule, alerts, unassigned = generator.generate()
            
            st.session_state.current_schedule = schedule
            st.session_state.alerts = alerts
            st.session_state.schedules[str(week_start)] = {
                'schedule': schedule,
                'alerts': alerts,
                'unassigned': unassigned,
                'opportunities': generator.opportunities_24hr
            }
            
            st.success("✅ Schedule generated successfully!")
            
            if alerts:
                st.subheader("📋 Alerts & Issues")
                for alert in alerts:
                    if alert['type'] == 'error':
                        st.error(alert['message'])
                    elif alert['type'] == 'warning':
                        st.warning(alert['message'])
                    else:
                        st.success(alert['message'])
            
            excel_file = export_to_excel(
                schedule,
                st.session_state.employees,
                st.session_state.sites,
                alerts,
                unassigned,
                generator.opportunities_24hr,
                week_start
            )
            
            st.download_button(
                label="📥 Download Excel Schedule",
                data=excel_file,
                file_name=f"rota_schedule_{week_start}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

def view_schedule():
    st.title("📅 View Schedule")
    
    if not st.session_state.schedules:
        st.info("No schedules generated yet. Go to 'Generate Schedule' to create your first rota!")
        return
    
    week_options = list(st.session_state.schedules.keys())
    selected_week = st.selectbox("Select Week", week_options)
    
    if selected_week:
        schedule_data = st.session_state.schedules[selected_week]
        schedule = schedule_data['schedule']
        
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        st.subheader(f"Week of {selected_week}")
        
        for emp in st.session_state.employees:
            if emp['id'] in schedule:
                with st.expander(f"👤 {emp['name']}", expanded=True):
                    cols = st.columns(7)
                    
                    for i, day in enumerate(days):
                        day_shifts = schedule[emp['id']].get(day, [])
                        
                        with cols[i]:
                            st.write(f"**{day[:3]}**")
                            if day_shifts:
                                shift = day_shifts[0]
                                st.write(f"🏢 {shift['site_name']}")
                                st.write(f"⏰ {shift['start']}-{shift['end']}")
                                st.write(f"📊 {shift['hours']:.1f}h")
                            else:
                                st.write("OFF")

if __name__ == "__main__":
    main()
